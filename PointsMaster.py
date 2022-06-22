import collections
import re
import openpyxl


class PointsMaster(object):

    def __init__(self, workbook, sheet, OUTPUT_FILE="errors.txt", DEVICE_ID=1, POINT_NAME=2, DEVICE_TYPE=3,
                 DESCRIPTION=5, AVAILABLE=8, start=20):
        self.workbook = workbook
        self.sheet = sheet
        self.start = start
        self.HEADER = 0
        self.DNP_INDEX = 6
        self.STATE_TABLE = 7
        self.UNITS = 7
        self.SUBSTATION = 0
        self.UNITS = 7
        self.POINT_NAME = POINT_NAME
        self.DEVICE_TYPE = DEVICE_TYPE
        self.DEVICE_ID = DEVICE_ID
        self.AVAILABLE = AVAILABLE
        self.DESCRIPTION = DESCRIPTION
        self.OUTPUT_FILE = OUTPUT_FILE
        self.error_count = 0

        # Delete old contents
        open(self.OUTPUT_FILE, 'w').close()

        wb = openpyxl.load_workbook('Attachment 3-Attribute Names & Engineering Units.xlsx', data_only=True)
        self.valid_points = set()
        self.valid_descriptions = set()

        self.breaker_relay_points = set()
        self.bus_relay_points = set()
        self.transformer_relay_points = set()
        self.line_relay_points = set()
        self.reactive_power_relay_points = set()
        self.meter_points = set()
        self.ppc_points = set()
        self.top_points = set()
        self.transformer_points = set()
        self.met_points = set()
        self.inverter_points = set()
        self.tracker_controller_points = set()
        self.soiling_station_points = set()

        self.breaker_relay_descriptions = set()
        self.bus_relay_descriptions = set()
        self.transformer_relay_descriptions = set()
        self.line_relay_descriptions = set()
        self.reactive_power_relay_descriptions = set()
        self.meter_descriptions = set()
        self.ppc_descriptions = set()
        self.top_descriptions = set()
        self.transformer_descriptions = set()
        self.met_descriptions = set()
        self.inverter_descriptions = set()
        self.tracker_controller_descriptions = set()
        self.soiling_station_descriptions = set()

        # point -> units
        self.breaker_relay_units = dict()
        self.bus_relay_units = dict()
        self.transformer_relay_units = dict()
        self.line_relay_units = dict()
        self.reactive_power_relay_units = dict()
        self.meter_units = dict()
        self.ppc_units = dict()
        self.top_units = dict()
        self.transformer_units = dict()
        self.met_units = dict()
        self.inverter_units = dict()
        self.tracker_controller_units = dict()
        self.soiling_station_units = dict()

        breaker_relay_points_sheet = wb['Breaker Relay']
        bus_relay_points_sheet = wb['Bus Relay']
        transformer_relay_points_sheet = wb['Transformer Relay']
        line_relay_points_sheet = wb['Line Relay']
        reactive_power_relay_points_sheet = wb['Reactive Power Relay']
        meter_points_sheet = wb['Meter']
        ppc_points_sheet = wb['PPC']
        top_points_sheet = wb['TOP']
        transformer_points_sheet = wb['Medium-High Voltage Transformer']
        met_points_sheet = wb['MET Station']
        inverter_points_sheet = wb['Inverter']
        tracker_controller_points_sheet = wb['Tracker Controller']
        soiling_station_points_sheet = wb['Soiling Station']

        for row in breaker_relay_points_sheet.values:
            if row[0] and row[0].strip() != 'Attribute Name':
                self.breaker_relay_points.add(row[0].strip())
            if row[3]:
                self.breaker_relay_descriptions.add(row[3].strip())
            if row[1] and 'binary' not in row[2].lower():
                self.breaker_relay_units[row[0]] = row[1]

        for row in bus_relay_points_sheet.values:
            if row[0] and row[0].strip() != 'Attribute Name':
                self.bus_relay_points.add(row[0].strip())
            if row[3]:
                self.bus_relay_descriptions.add(row[3].strip())
            if row[1] and 'binary' not in row[2].lower():
                self.bus_relay_units[row[0]] = row[1]

        for row in transformer_relay_points_sheet.values:
            if row[0] and row[0].strip() != 'Attribute Name':
                self.transformer_relay_points.add(row[0].strip())
            if row[3]:
                self.transformer_relay_descriptions.add(row[3].strip())
            if row[1] and 'binary' not in row[2].lower():
                self.transformer_relay_units[row[0]] = row[1]

        for row in line_relay_points_sheet.values:
            if row[0] and row[0].strip() != 'Attribute Name':
                self.line_relay_points.add(row[0].strip())
            if row[3]:
                self.line_relay_descriptions.add(row[3].strip())
            if row[1] and 'binary' not in row[2].lower():
                self.line_relay_units[row[0]] = row[1]

        for row in reactive_power_relay_points_sheet.values:
            if row[0] and row[0].strip() != 'Attribute Name':
                self.reactive_power_relay_points.add(row[0].strip())
            if row[3]:
                self.reactive_power_relay_descriptions.add(row[3].strip())
            if row[1] and 'binary' not in row[2].lower():
                self.reactive_power_relay_units[row[0]] = row[1]

        for row in meter_points_sheet.values:
            if row[0] and row[0].strip() != 'Attribute Name':
                self.meter_points.add(row[0].strip())
            if row[3]:
                self.meter_descriptions.add(row[3].strip())
            if row[1] and 'binary' not in row[2].lower():
                self.meter_units[row[0]] = row[1]

        for row in ppc_points_sheet.values:
            if row[0] and row[0].strip() != 'Attribute Name':
                self.ppc_points.add(row[0].strip())
            if row[3]:
                self.ppc_descriptions.add(row[3].strip())
            if row[1] and 'binary' not in row[2].lower():
                self.ppc_units[row[0]] = row[1]

        for row in top_points_sheet.values:
            if row[0] and row[0].strip() != 'Attribute Name':
                self.top_points.add(row[0].strip())
            if row[3]:
                self.top_descriptions.add(row[3].strip())
            if row[1] and 'binary' not in row[2].lower():
                self.top_units[row[0]] = row[1]

        for row in transformer_points_sheet.values:
            if row[0] and row[0].strip() != 'Attribute Name':
                self.transformer_points.add(row[0].strip())
            if row[3]:
                self.transformer_descriptions.add(row[3].strip())
            if row[1] and 'binary' not in row[2].lower():
                self.transformer_units[row[0]] = row[1]

        for row in met_points_sheet.values:
            if row[0] and row[0].strip() != 'Attribute Name':
                self.met_points.add(row[0].strip())
            if row[3]:
                self.met_descriptions.add(row[3].strip())
            if row[1] and 'binary' not in row[2].lower():
                self.met_units[row[0]] = row[1]

        for row in inverter_points_sheet.values:
            if row[0] and row[0].strip() != 'Attribute Name':
                self.inverter_points.add(row[0].strip())
            if row[3]:
                self.inverter_descriptions.add(row[3].strip())
            if row[1] and 'binary' not in row[2].lower():
                self.inverter_units[row[0]] = row[1]

        for row in tracker_controller_points_sheet.values:
            if row[0] and row[0].strip() != 'Attribute Name':
                self.tracker_controller_points.add(row[0].strip())
            if row[3]:
                self.tracker_controller_descriptions.add(row[3].strip())
            if row[1] and 'binary' not in row[2].lower():
                self.tracker_controller_units[row[0]] = row[1]

        for row in soiling_station_points_sheet.values:
            if row[0] and row[0].strip() != 'Attribute Name':
                self.soiling_station_points.add(row[0].strip())
            if row[3]:
                self.soiling_station_descriptions.add(row[3].strip())
            if row[1] and 'binary' not in row[2].lower():
                self.soiling_station_units[row[0]] = row[1]

    def validate_point_length(self):
        with open(self.OUTPUT_FILE, mode="a") as f:
            wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
            for i, row in enumerate(wb[self.sheet]):
                point_name = str(row[self.POINT_NAME].value).strip()
                is_available = str(row[self.AVAILABLE].value).strip()
                if i > self.start and point_name != 'None' and point_name != 'Point Name' \
                        and is_available == "Not Requested-Available":
                    name_length = len(point_name)
                    if name_length > 60:
                        self.error_count += 1
                        f.write("Row {} - {} - Exceeds 60 char limit by {} chars\n".
                                format(i + 1, point_name, name_length - 60))
                        print(("Row {} - {} - Exceeds 60 char limit by {} chars".
                               format(i + 1, point_name, name_length - 60)))

    def validate_underscore_usage(self):
        with open(self.OUTPUT_FILE, mode="a") as f:
            wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
            for i, row in enumerate(wb[self.sheet]):
                point_name = str(row[self.POINT_NAME].value).strip()
                if i > self.start and point_name != 'None' and point_name != 'Point Name':
                    number_of_underscores = point_name.count("_")
                    if number_of_underscores != 3:
                        self.error_count += 1
                        f.write("Row {} - {} - has {} underscores when there should be exactly 3 per A11\n".
                                format(i + 1, point_name, number_of_underscores))
                        print("Row {} - {} - has {} underscores when there should be exactly 3 per A11".
                              format(i + 1, point_name, number_of_underscores))

    def validate_point_names(self):
        wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
        with open(self.OUTPUT_FILE, mode="a") as f:
            for i, row in enumerate(wb[self.sheet]):
                point_name = str(row[self.POINT_NAME].value).strip()
                is_available = str(row[self.AVAILABLE].value).strip()
                device_type = str(row[self.DEVICE_TYPE].value).strip()
                if i > self.start and point_name != 'None' and point_name != 'Point Name' \
                        and is_available != "Not Requested-Available":
                    if device_type == 'Meter':
                        valid_names = self.meter_points.copy()
                    elif device_type == 'Breaker Relay':
                        valid_names = self.breaker_relay_points.copy()
                    elif device_type == 'Bus Relay':
                        valid_names = self.bus_relay_points.copy()
                    elif device_type == 'Transformer Relay':
                        valid_names = self.transformer_relay_points.copy()
                    elif device_type == 'Line Relay':
                        valid_names = self.line_relay_points.copy()
                    elif device_type == 'Reactive Power Relay':
                        valid_names = self.reactive_power_relay_points.copy()
                    elif device_type == 'MET Station':
                        valid_names = self.met_points.copy()
                    elif device_type == "PPC" or device_type == "Power Plant Controller":
                        valid_names = self.ppc_points.copy()
                    elif device_type == "TOP":
                        valid_names = self.top_points.copy()
                    elif device_type == 'Medium-High Voltage Transformer':
                        valid_names = self.transformer_points.copy()
                    elif device_type == 'Inverter':
                        valid_names = self.inverter_points.copy()
                    elif device_type == 'Tracker Controller':
                        valid_names = self.tracker_controller_points.copy()
                    elif device_type == 'Soiling Station':
                        valid_names = self.soiling_station_points.copy()
                    elif device_type == 'Discrete I/O':
                        continue  # No defined names for this device type
                    elif device_type == 'Analog I/O':
                        continue  # No defined names for this device type
                    elif device_type == 'Annunciator':
                        continue  # No defined names for this device type
                    else:
                        self.error_count += 1
                        f.write("Row {} - Source Device Type {} invalid\n".format(i + 1, device_type))
                        print("Row {} - Source Device Type {} invalid".format(i + 1, device_type))
                        continue

                    if not self._point_is_valid(point_name, valid_names):
                        self.error_count += 1
                        print("Row {} - {} - Invalid Point Name".format(i + 1, point_name))
                        f.write("Row {} - {} - Invalid Point Name\n".format(i + 1, point_name))

    def _point_is_valid(self, name, valid_names) -> bool:
        for valid_name in valid_names:
            if valid_name in name:
                return True
            if '#' in valid_name:
                string_split = valid_name.split("#")
                part1 = string_split[0]
                part2 = string_split[1]
                if part1 in name and part2 in name:
                    return True
        return False

    def verify_availability_entry_is_valid(self):
        wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
        with open(self.OUTPUT_FILE, mode="a") as f:
            for i, row in enumerate(wb[self.sheet]):
                availability = str(row[self.AVAILABLE].value).strip()
                if availability != "Requested-Not Available" and availability != "Requested-Available" and \
                        availability != "Requested-Not Applicable" and availability != "Not Requested-Available":
                    self.error_count += 1
                    print("Row {} - {} - Invalid Availability. Valid options are Requested-Not Available, "
                          "Requested-Available, Requested-Not Applicable, Not Requested-Available"
                          .format(i + 1, availability))
                    f.write("Row {} - {} - Invalid Availability. Valid options are Requested-Not Available, "
                            "Requested-Available, Requested-Not Applicable, Not Requested-Available\n"
                            .format(i + 1, availability))

    def verify_units_are_correct(self):
        wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
        has_units = False
        with open(self.OUTPUT_FILE, mode="a") as f:
            for i, row in enumerate(wb[self.sheet]):
                point_type_header = str(row[self.HEADER].value).strip()
                if point_type_header == "Digital Inputs" or point_type_header == "Digital Outputs":
                    has_units = False
                elif point_type_header == "Analog Inputs" or point_type_header == "Counters" or point_type_header == "Analog Outputs":
                    has_units = True
                if has_units:
                    point_name = str(row[self.POINT_NAME].value).strip()
                    try:
                        point_attribute = point_name.split("_")[-1]
                    except:
                        # Other check will catch if they don't use underscores but don't want to crash
                        continue
                    units = str(row[self.UNITS].value).strip()
                    device_type = str(row[self.DEVICE_TYPE].value).strip()
                    correct_unit = None

                    if device_type == 'Meter':
                        correct_unit = self.meter_units.get(point_attribute)
                    elif device_type == 'Breaker Relay':
                        correct_unit = self.breaker_relay_units.get(point_attribute)
                    elif device_type == 'Bus Relay':
                        correct_unit = self.bus_relay_units.get(point_attribute)
                    elif device_type == 'Transformer Relay':
                        correct_unit = self.transformer_relay_units.get(point_attribute)
                    elif device_type == 'Line Relay':
                        correct_unit = self.line_relay_units.get(point_attribute)
                    elif device_type == 'Reactive Power Relay':
                        correct_unit = self.reactive_power_relay_units.get(point_attribute)
                    elif device_type == 'MET Station':
                        correct_unit = self.met_units.get(point_attribute)
                    elif device_type == "PPC" or device_type == "Power Plant Controller":
                        correct_unit = self.ppc_units.get(point_attribute)
                    elif device_type == "TOP":
                        correct_unit = self.top_units.get(point_attribute)
                    elif device_type == 'Medium-High Voltage Transformer':
                        correct_unit = self.transformer_units.get(point_attribute)
                    elif device_type == 'Inverter':
                        correct_unit = self.inverter_units.get(point_attribute)
                    elif device_type == 'Tracker Controller':
                        correct_unit = self.tracker_controller_units.get(point_attribute)
                    elif device_type == 'Soiling Station':
                        correct_unit = self.soiling_station_units.get(point_attribute)

                    if correct_unit and correct_unit != units:
                        self.error_count += 1
                        print("Row {} - {} - Invalid Unit, should be {}".format(i + 1, units, correct_unit))
                        f.write("Row {} - {} - Invalid Unit should be {}\n".format(i + 1, units, correct_unit))

    def validate_point_descriptions(self):
        wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
        with open(self.OUTPUT_FILE, mode="a") as f:
            for i, row in enumerate(wb[self.sheet]):
                point_name = str(row[self.POINT_NAME].value).strip()
                is_available = str(row[self.AVAILABLE].value).strip()
                device_type = str(row[self.DEVICE_TYPE].value).strip()
                description = str(row[self.DESCRIPTION].value).strip()
                if i > self.start and point_name != 'None' and point_name != 'Point Name' \
                        and is_available != "Not Requested-Available":
                    if device_type == 'Meter':
                        descriptions = self.meter_descriptions.copy()
                    elif device_type == 'Breaker Relay':
                        descriptions = self.breaker_relay_descriptions.copy()
                    elif device_type == 'Bus Relay':
                        descriptions = self.bus_relay_descriptions.copy()
                    elif device_type == 'Transformer Relay':
                        descriptions = self.transformer_relay_descriptions.copy()
                    elif device_type == 'Line Relay':
                        descriptions = self.line_relay_descriptions.copy()
                    elif device_type == 'Reactive Power Relay':
                        descriptions = self.reactive_power_relay_descriptions.copy()
                    elif device_type == 'MET Station':
                        descriptions = self.met_descriptions.copy()
                    elif device_type == "PPC" or device_type == "Power Plant Controller":
                        descriptions = self.ppc_descriptions.copy()
                    elif device_type == "TOP":
                        descriptions = self.top_descriptions.copy()
                    elif device_type == 'Medium-High Voltage Transformer':
                        descriptions = self.transformer_descriptions.copy()
                    elif device_type == 'Inverter':
                        descriptions = self.inverter_descriptions.copy()
                    elif device_type == 'Tracker Controller':
                        descriptions = self.tracker_controller_descriptions.copy()
                    elif device_type == 'Soiling Station':
                        descriptions = self.soiling_station_descriptions.copy()
                    elif device_type == 'Discrete I/O':
                        continue  # No defined descriptions for these
                    elif device_type == 'Analog I/O':
                        continue  # No defined descriptions for these
                    elif device_type == 'Annunciator':
                        continue  # No defined descriptions for these
                    else:
                        continue

                    if description not in descriptions:
                        self.error_count += 1
                        print("Row {} - {} - Invalid Description".format(i + 1, description))
                        f.write("Row {} - {} - Invalid Description\n".format(i + 1, description))

    def verify_device_types_have_all_points(self):
        device_dictionary = {}
        wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
        with open(self.OUTPUT_FILE, mode="a") as f:
            for i, row in enumerate(wb[self.sheet]):
                point_name = str(row[self.POINT_NAME].value).strip()
                device_type = str(row[self.DEVICE_TYPE].value).strip()
                source_device = str(row[self.DEVICE_ID].value).strip()
                if i > self.start and point_name != 'None' and point_name != 'Point Name':
                    if source_device not in device_dictionary:
                        if device_type == 'Meter':
                            points = self.meter_points.copy()
                        elif device_type == 'Breaker Relay':
                            points = self.breaker_relay_points.copy()
                        elif device_type == 'Bus Relay':
                            points = self.bus_relay_points.copy()
                        elif device_type == 'Transformer Relay':
                            points = self.transformer_relay_points.copy()
                        elif device_type == 'Line Relay':
                            points = self.line_relay_points.copy()
                        elif device_type == 'Reactive Power Relay':
                            points = self.reactive_power_relay_points.copy()
                        elif device_type == 'MET Station':
                            points = self.met_points.copy()
                        elif device_type == "PPC" or device_type == "Power Plant Controller":
                            points = self.ppc_points.copy()
                        elif device_type == "TOP":
                            points = self.top_points.copy()
                        elif device_type == 'Medium-High Voltage Transformer':
                            points = self.transformer_points.copy()
                        elif device_type == 'Inverter':
                            points = self.inverter_points.copy()
                        elif device_type == 'Tracker Controller':
                            points = self.tracker_controller_points.copy()
                        elif device_type == 'Soiling Station':
                            points = self.soiling_station_points.copy()
                        elif device_type == 'Discrete I/O':
                            continue  # No defined set of points for this device type
                        elif device_type == 'Analog I/O':
                            continue  # No defined set of points for this device type
                        elif device_type == 'Annunciator':
                            continue  # No defined set of points for this device type
                        else:
                            continue
                        device_dictionary[source_device] = points

                    dictionary_pts = device_dictionary.get(source_device)
                    for pt in dictionary_pts.copy():  # Have to copy or RuntimeError: Set changed size during iteration
                        if '#' in pt:
                            string_split = pt.split("#")
                            part1 = string_split[0]
                            part2 = string_split[1]
                            if part1 in point_name and part2 in point_name:
                                dictionary_pts.remove(pt)
                        elif pt in point_name:
                            dictionary_pts.remove(pt)

            for source_dev, remaining_pts in device_dictionary.items():
                for remaining_pt in remaining_pts:
                    self.error_count += 1
                    f.write("Missing point for {}, {}\n".format(source_dev, remaining_pt))
                    print("Missing point for {}, {}".format(source_dev, remaining_pt))

    def validate_all_placeholders_are_removed(self):
        with open(self.OUTPUT_FILE, mode="a") as f:
            wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
            for i, row in enumerate(wb[self.sheet]):
                point_name = str(row[self.POINT_NAME].value).strip()
                availability = str(row[self.AVAILABLE].value).strip()
                if i > self.start and point_name != 'None' and point_name != 'Point Name':
                    if '#' in point_name and availability != "Requested-Not Available":
                        self.error_count += 1
                        print("Row {} - Placeholder left in point {}".format(i + 1, point_name))
                        f.write("Row {} - Placeholder left in point {}\n".format(i + 1, point_name))

    def get_error_count(self):
        return self.error_count

    def check_for_duplicates(self):
        device_dictionary = {}
        wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
        with open(self.OUTPUT_FILE, mode="a") as f:
            for i, row in enumerate(wb[self.sheet]):
                point_name = str(row[self.POINT_NAME].value).strip()
                source_device = str(row[self.DEVICE_ID].value).strip()
                if i > self.start and point_name != 'None' and point_name != 'Point Name':
                    if source_device not in device_dictionary:
                        device_dictionary[source_device] = [point_name]
                    else:
                        device_dictionary[source_device].append(point_name)

            for source_dev, points in device_dictionary.items():
                duplicates = [item for item, count in collections.Counter(points).items() if count > 1]
                for duplicate in duplicates:
                    self.error_count += 1
                    f.write("Duplicate point for {}, {}\n".format(source_dev, duplicate))
                    print("Duplicate point for {}, {}\n".format(source_dev, duplicate))

    def verify_device_ids(self):
        wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
        with open(self.OUTPUT_FILE, mode="a") as f:
            for i, row in enumerate(wb[self.sheet]):
                point_name = str(row[self.POINT_NAME].value).strip()
                device_type = str(row[self.DEVICE_TYPE].value).strip()
                device_id = str(row[self.DEVICE_ID].value).strip()
                if i > self.start and point_name.count("_") == 3:
                    if device_type == 'Meter':
                        pattern = "[A-Z\-\d]* HS|[A-Z\-\d]* LS"
                    elif device_type == 'Breaker Relay':
                        pattern = "[A-Z\-\d]*"
                    elif device_type == 'Bus Relay':
                        pattern = "[A-Z\-\d]*"
                    elif device_type == 'Transformer Relay':
                        pattern = "[A-Z\-\d]*"
                    elif device_type == 'Line Relay':
                        pattern = "[A-Z\-\d]*"
                    elif device_type == 'Reactive Power Relay':
                        pattern = "[A-Z\-\d]*"
                    elif device_type == 'MET Station':
                        pattern = "MET\d{1,2}"
                    elif device_type == "PPC" or device_type == "Power Plant Controller":
                        pattern = "PPC"
                    elif device_type == "TOP":
                        pattern = "TOP"
                    elif device_type == 'Medium-High Voltage Transformer':
                        pattern = "XFMR\d{3}"
                    elif device_type == 'Inverter':
                        pattern = "INV\d{3}"
                    elif device_type == 'Tracker Controller':
                        pattern = "TC\d{3}\-\d{2}_M\d{3}"
                    elif device_type == 'Soiling Station':
                        pattern = "SOIL\d{3}"
                    elif device_type == 'Discrete I/O':
                        pattern = "[A-Z\-\d]*"
                    elif device_type == 'Analog I/O':
                        pattern = "[A-Z\-\d]*"
                    elif device_type == 'Annunciator':
                        pattern = "[A-Z\-\d]*"
                    else:
                        pattern = None
                        print("Row {} contains Device Type {} which is not a valid option".format(i + 1, device_type))
                        f.write("Row {} contains Device Type {} which is not a valid option\n".format(i + 1, device_type))
                        self.error_count += 1

                    # Verify all Device IDs are capitalized
                    for letter in device_id:
                        if letter.islower():
                            print("Row {} Lowercase characters not allowed in DeviceID {}".format(i + 1, device_id))
                            f.write("Row {} Lowercase characters not allowed in DeviceID {}\n".format(i + 1, device_id))
                            self.error_count += 1
                            break

                    if pattern:
                        # Verify that the device id was used in the point name
                        if device_id not in point_name:
                            print("Row {} Device ID {} not used in Point name".format(i + 1, device_id))
                            f.write("Row {} Device ID {} not used in Point name\n".format(i + 1, device_id))
                            self.error_count += 1

                        # Verify that Device ID follows the Clearway convention
                        match = re.search(pattern, device_id)
                        if not match:
                            print("Row {} Device ID {} does not follow A11 convention for Device Type {}".format(i + 1, device_id, device_type))
                            f.write("Row {} Device ID {} does not follow A11 convention for Device Type {}\n".format(i + 1, device_id, device_type))
                            self.error_count += 1

    def verify_substation_name_in_points(self):
        wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
        with open(self.OUTPUT_FILE, mode="a") as f:
            for i, row in enumerate(wb[self.sheet]):
                if i > self.start:
                    point_name = str(row[self.POINT_NAME].value).strip()
                    substation = str(row[self.SUBSTATION].value).strip()
                    substation_with_underscores = "_" + substation + "_"
                    if substation_with_underscores not in point_name and point_name != 'None' and point_name and point_name != 'Point Name':
                        print("Row {} Point Name does not include substation {}".format(i + 1, point_name, substation))
                        f.write("Row {} Point Name does not include substation {}\n".format(i + 1, point_name, substation))
                        self.error_count += 1

    def validate_dnp_indexes_not_reused(self):
        analog_input_index_dictionary = {}
        binary_input_index_dictionary = {}
        counters_index_dictionary = {}
        analog_output_index_dictionary = {}
        binary_output_index_dictionary = {}
        active_dictionary = None
        wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
        with open(self.OUTPUT_FILE, mode="a") as f:
            for i, row in enumerate(wb[self.sheet]):
                point_type_header = str(row[self.HEADER].value).strip()
                if point_type_header == "Analog Inputs":
                    active_dictionary = analog_input_index_dictionary
                elif point_type_header == "Digital Inputs":
                    active_dictionary = binary_input_index_dictionary
                elif point_type_header == "Counters":
                    active_dictionary = counters_index_dictionary
                elif point_type_header == "Analog Outputs":
                    active_dictionary = analog_output_index_dictionary
                elif point_type_header == "Digital Outputs":
                    active_dictionary = binary_output_index_dictionary

                dnp_index = row[self.DNP_INDEX].value
                if type(dnp_index) == int and active_dictionary is not None:
                    row = active_dictionary.get(dnp_index)
                    if row:
                        print("Row {} used the same DNP index as row {}".format(i+1, row))
                        f.write("Row {} used the same DNP index as row {}\n".format(i+1, row))
                        self.error_count += 1
                    else:
                        active_dictionary[dnp_index] = i + 1

    def validate_only_available_points_have_dnp_indexes(self):
        wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
        with open(self.OUTPUT_FILE, mode="a") as f:
            for i, row in enumerate(wb[self.sheet]):
                dnp_index = row[self.DNP_INDEX].value
                is_available = str(row[self.AVAILABLE].value).strip().lower()
                if type(dnp_index) == int and (("not available" in is_available) or ("not applicable" in is_available)):
                    print("Row {} should not get a dnp index since it is not available/applicable".format(i+1))
                    f.write("Row {} should not get a dnp index since it is not available/applicable\n".format(i+1))
                    self.error_count += 1

    def find_obvious_state_table_errors(self):
        wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
        is_binary_input = False
        with open(self.OUTPUT_FILE, mode="a") as f:
            for i, row in enumerate(wb[self.sheet]):
                point_type_header = str(row[self.HEADER].value).strip()
                if point_type_header == "Digital Inputs":
                    is_binary_input = True
                elif point_type_header == "Analog Inputs" or point_type_header == "Counters" or point_type_header == "Digital Outputs" or point_type_header == "Analog Outputs":
                    is_binary_input = False
                if is_binary_input:
                    point_name = str(row[self.POINT_NAME].value).strip().lower()
                    state_table = str(row[self.STATE_TABLE].value).strip().lower()
                    if "remote" in point_name and "local" in point_name:
                        if "remote" not in state_table or "local" not in state_table:
                            print("Row {} 0/1 State is incorrect, needs to be 0=Remote, 1=Local or 1=Remote, 0=Local".format(i+1))
                            f.write("Row {} 0/1 State is incorrect, needs to be 0=Remote, 1=Local or 1=Remote, 0=Local\n".format(i+1))
                            self.error_count += 1
                    if "alarm" in point_name and "alarm" not in state_table:
                        print("Row {} is an alarm point, 0/1 state should be 0=Normal, 1=Alarm or 1=Normal, 0=Alarm".format(i+1))
                        f.write("Row {} is an alarm point, 0/1 state should be 0=Normal, 1=Alarm or 1=Normal, 0=Alarm\n".format(i+1))
                        self.error_count += 1
                    if "breaker status" in point_name:
                        if "open" not in state_table or "closed" not in state_table:
                            print("Row {} 0/1 State is incorrect, needs to be 0=Open, 1=Closed".format(i+1))
                            f.write("Row {} 0/1 State is incorrect, needs to be 0=Open, 1=Closed\n".format(i+1))
                            self.error_count += 1
                    if "switch status" in point_name:
                        if "open" not in state_table or "closed" not in state_table:
                            print("Row {} 0/1 State is incorrect, needs to be 0=Open, 1=Closed".format(i+1))
                            f.write("Row {} 0/1 State is incorrect, needs to be 0=Open, 1=Closed\n".format(i+1))
                            self.error_count += 1
                    if "breaker fail initiate" in point_name:
                        if "normal" not in state_table or "bfi" not in state_table:
                            print("Row {} 0/1 State is incorrect, needs to be 0=Normal, 1=BFI".format(i+1))
                            f.write("Row {} 0/1 State is incorrect, needs to be 0=Normal, 1=BFI\n".format(i+1))
                            self.error_count += 1
                    if "line comm cutout" in point_name:
                        if "normal" not in state_table or "cutout" not in state_table:
                            print("Row {} 0/1 State is incorrect, needs to be 0=Normal, 1=Cutout".format(i+1))
                            f.write("Row {} 0/1 State is incorrect, needs to be 0=Normal, 1=Cutout\n".format(i+1))
                            self.error_count += 1
                    if "communication status" in point_name:
                        if "offline" not in state_table or "online" not in state_table:
                            print("Row {} 0/1 State is incorrect, needs to be 0=Offline, 1=Online".format(i+1))
                            f.write("Row {} 0/1 State is incorrect, needs to be 0=Offline, 1=Online\n".format(i+1))
                            self.error_count += 1
                    if "relay enabled" in point_name:
                        if "disabled" not in state_table or "enabled" not in state_table:
                            print("Row {} 0/1 State is incorrect, needs to be 0=Disabled, 1=Enabled".format(i+1))
                            f.write("Row {} 0/1 State is incorrect, needs to be 0=Disabled, 1=Enabled\n".format(i+1))
                            self.error_count += 1

    def verify_suggested_names_for_not_requested_points_follow_pascal_case(self):
        wb = openpyxl.load_workbook(self.workbook, data_only=True, read_only=True)
        with open(self.OUTPUT_FILE, mode="a") as f:
            for i, row in enumerate(wb[self.sheet]):
                point_name = str(row[self.POINT_NAME].value).strip()
                availability = str(row[self.AVAILABLE].value).strip()
                if availability == "Not Requested-Available" and point_name.count('_') == 3:  # Other checks ensure this
                    attribute_name = re.sub(".*_.*_.*_(.*)", r"\1", point_name)
                    attribute_words = attribute_name.split(" ")
                    if self.__has_invalid_casing(attribute_words):
                        print("Row {} Error - {} casing invalid. Not Requested, Available Point names must be Pascal case with spaces between words. i.e. CSCSB1 Switch Control".format(i+1, attribute_name))
                        f.write("Row {} Error - {} casing invalid. Not Requested, Available Point names must be Pascal case with spaces between words. i.e. CSCSB1 Switch Control\n".format(i+1, attribute_name))
                        self.error_count += 1

    def __has_invalid_casing(self, words) -> bool:
        is_all_caps = True
        for word in words:
            for i in range(len(word)):
                if i == 0 and word[i].islower():
                    return True
                elif i != 0 and word[i].islower():
                    is_all_caps = False
        return True if is_all_caps else False


